# Author: Karthik Reddy Pagilla

import random
import numpy as np
from copy import deepcopy
import datetime
from openpyxl import Workbook

global nodes_visited
global goal_state

def state_generator():
    random_numbers = random.sample(range(9), 9)
    state = np.reshape(random_numbers, (3,3))

    return state

def goal_check(state):
    global goal_state

    flat_state = list(state.flatten())
    flat_goal = list(goal_state.flatten())

    for i in range(0, 9):
        if flat_state[i] != flat_goal[i]:
            return False
    
    return True

class Node:
    def __init__(self, state):
        self.state = state
        self.children = []
        self.parent = None
        self.depth = 0
        self.path_cost = 0

    def toString():
        print("The Path Cost: " + str(self.path_cost))

def getEmptyPosition(state):
    for i in range(0, len(state)):
        for j in range(0, len(state[i])):
            if state[i][j] == 0:
                return (i, j)

    return None

def is_solvable(state):
    list_state = list(state.flatten())
    inversions = 0

    for i in range(0, len(list_state)):
        for j in range(i + 1, len(list_state)):
            if ((list_state[i] != 0) and (list_state[j] != 0) and (list_state[i] > list_state[j])):
                inversions += 1

    if inversions % 2 == 0:
        return "True"

    return "False"

def heuristic(initial_state, hType):
    displaced = 0
    distance = 0

    global goal_state

    if hType == "Displaced Tiles":
        initial_flat = list(initial_state.flatten())
        goal_flat = list(goal_state.flatten())

        for i in range(0, 9):
            if initial_flat[i] != goal_flat[i] and initial_flat[i] != 0:
                displaced += 1
        return displaced
    elif hType == "Manhattan Distance":
        ini_cords = cord_generator(initial_state)
        goal_cords = cord_generator(goal_state)

        for i in range(1, 9):
            (x1, y1) = ini_cords[i]
            (x2, y2) = goal_cords[i]

            x_change = x1 - x2
            y_change = y1 - y2

            distance += abs(x_change) + abs(y_change)

        return distance

    return None

def cord_generator(state):
    cords = {}

    for x, row in enumerate(state):
        for y, val in enumerate(row):
            cords[val] = (x, y)
    
    return cords

def north(curr_state):
    (x, y) = getEmptyPosition(curr_state)

    state = deepcopy(curr_state)

    temp = state[x - 1][y]
    state[x - 1][y] = 0
    state[x][y] = temp

    return state

def south(curr_state):
    (x, y) = getEmptyPosition(curr_state)
    
    state = deepcopy(curr_state)

    temp = state[x + 1][y]
    state[x + 1][y] = 0
    state[x][y] = temp

    return state

def east(curr_state):
    (x, y) = getEmptyPosition(curr_state)

    state = deepcopy(curr_state)

    temp = state[x][y + 1]
    state[x][y + 1] = 0
    state[x][y] = temp

    return state

def west(curr_state):
    (x, y) = getEmptyPosition(curr_state)

    state = deepcopy(curr_state)

    temp = state[x][y - 1]
    state[x][y - 1] = 0
    state[x][y] = temp

    return state

def expand_node(node):
    global nodes_visited

    moves = []
    state = node.state

    (x, y) = getEmptyPosition(state)

    if (x != 0):
        moves.append("north")

    if (x != 2):
        moves.append("south")

    if (y != 0):
        moves.append("west")

    if (y != 2):
        moves.append("east")

    for move in moves:
        if (move == "north"):
            c = north(state)
        elif (move == "south"):
            c = south(state)
        elif (move == "west"):
            c = west(state)
        elif (move == "east"):
            c = east(state)

        child = Node(c)
        child.parent = node
        child.depth = node.depth + 1
        child.path_cost = node.path_cost + 1

        node.children.append(child)

    nodes_visited += 1

    return node

def evaluate_node(node, hType):
    return node.path_cost + heuristic(node.state, hType)

def node_to_expand(fringe, hType):
    min_value = 10000000
    expand_node = None
    for node in fringe:
        if evaluate_node(node, hType) < min_value:
            min_value = evaluate_node(node, hType)
            expand_node = node
    
    return expand_node

def string_maker(state):
    flat_state = list(state.flatten())

    res = ""

    for i in flat_state:
        res += str(i)

    return res

def general_search(hType):
    global nodes_visited
    global goal_state

    nodes_visited = 0

    # state = np.array([[0,2,3],[1,4,5],[8,7,6]])
    # goal_state = np.array([[1,2,3],[8,0,4],[7,6,5]])

    state = state_generator()
    goal_state = state_generator()

    if is_solvable(state) != is_solvable(goal_state):
        return "Not Solvable"


    node = Node(state)
    node.path_cost = 0
    node.depth = 0
    node.parent = None

    fringe = []
    closed = []
    
    time = 0
    a = datetime.datetime.now()
    while(node != None and time < 150):

        if (goal_check(node.state)):
            return node

        node = expand_node(node)

        for child in node.children:
            if string_maker(child.state) not in closed:
                fringe.append(child)

        node = node_to_expand(fringe, hType)

        closed.append(string_maker(node.state))
        fringe.remove(node)
        b = datetime.datetime.now()

        diff = b - a
        time = diff.total_seconds()

    return "Not Solvable"

def path_finder(node):
    path = []
    while(node != None):
        path.append(string_maker(node.state))
        node = node.parent

    path.reverse()
    result = ""
    for state in path:
        if state == string_maker(goal_state):
            result = result + state
        else:
            result = result + state + " --> "
    
    return result

k = 1

workbook = Workbook()
sheet = workbook.active

for i in range(0, 5):
    print(k)

    a = datetime.datetime.now()
    result = general_search("Displaced Tiles")
    b = datetime.datetime.now()

    diff = b - a
    cpu_time = diff.total_seconds() * 1000

    c1 = sheet.cell(row = k, column = 1)
    c1.value = nodes_visited

    c2 = sheet.cell(row = k, column = 2)
    c3 = sheet.cell(row = k, column = 3)
    c4 = sheet.cell(row = k, column = 4)
    if result == "Not Solvable":
        c2.value = 0
        c4.value = "Not Solvable"
    else:
        c2.value = result.path_cost
        c4.value = path_finder(result)

    c3.value = cpu_time

    a1 = datetime.datetime.now()
    result1 = general_search("Manhattan Distance")
    b1 = datetime.datetime.now()

    diff1 = b1 - a1
    cpu_time1 = diff1.total_seconds() * 1000

    c5 = sheet.cell(row = k, column = 5)
    c5.value = nodes_visited

    c6 = sheet.cell(row = k, column = 6)
    c7 = sheet.cell(row = k, column = 7)
    c8 = sheet.cell(row = k, column = 8)
    if result1 == "Not Solvable":
        c6.value = 0
        c8.value = "Not Solvable"
    else:
        c6.value = result1.path_cost
        c8.value = path_finder(result1)

    c7.value = cpu_time1


    k += 1

workbook.save("/content/puzzle-complete.xlsx")
print("Done")